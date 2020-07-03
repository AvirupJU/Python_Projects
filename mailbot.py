import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('Liam.xlsx')
sheet = wb['Sheet1']

lastCol = 5

mybois = {}
mygals = {}
for r in range(2, 85):
    gender = sheet.cell(row=r, column=lastCol).value
    if gender == 'M':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        mybois[name] = email
    else :
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        mygals[name] = email
        
#print(mybois)
#print(mygals)

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('themail@gmail.com', 'passwords')

for name, email in mygals.items():
    body = "Hi, %s. As our campus has been severely damaged we have to postpone our training programme indefinitely. " \
           "In the meantime you can try touching your nose with your tongue and practise some tongue twisters." \
           "We'll let you know as soon as situations improve. Also there has been a change in the chain of command," \
           "Ms. Royal Blue will be taking charge of the training programme. Thank You." % name
    print("Sending email to %s..." % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

'''
for name, email in mybois.items():
    body = "Hi "
    print("Sending email to %s..." % email)
    sendmailStatus = smtpObj.sendmail('rebelsapienorg@gmail.com', email, body)

    if sendmailStatus !={}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
'''
        
smtpObj.quit()

